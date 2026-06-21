"""
Create the standard Excel parameter template for isolation bearing reports.
3 sheets: 项目基本信息, 产品型号清单, 检测数据
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def style_header(ws, row, cols, fill_color='B3D9FF'):
    """Apply header styling."""
    fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
    font = Font(name='微软雅黑', size=11, bold=True)
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal='center', vertical='center')


def style_data_rows(ws, start_row, end_row, cols):
    """Apply data row styling."""
    font = Font(name='微软雅黑', size=10)
    for row in range(start_row, end_row + 1):
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = font
            cell.alignment = Alignment(horizontal='center', vertical='center')


def auto_width(ws, cols):
    """Set reasonable column widths."""
    for col in range(1, cols + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18


def create_excel_template(output_path):
    wb = Workbook()

    # ===== Sheet 1: 项目基本信息 =====
    ws1 = wb.active
    ws1.title = '项目基本信息'

    headers1 = ['项目名称', '产品供应商', '制造地址', '联系电话', '制造日期']
    for col, h in enumerate(headers1, 1):
        ws1.cell(row=1, column=col, value=h)
    style_header(ws1, 1, len(headers1))

    # Sample data
    sample1 = ['北京吉利学院整体搬迁成都项目科学与艺术中心-1号楼项目',
               '四川融海运通抗震科技有限责任公司',
               '四川省成都市新津区普兴街道清云南路256号',
               '18180838170',
               '2024年12月']
    for col, val in enumerate(sample1, 1):
        ws1.cell(row=2, column=col, value=val)

    style_data_rows(ws1, 2, 3, len(headers1))
    auto_width(ws1, len(headers1))
    ws1.column_dimensions['A'].width = 45
    ws1.column_dimensions['B'].width = 35
    ws1.column_dimensions['C'].width = 40

    # ===== Sheet 2: 产品型号清单 =====
    ws2 = wb.create_sheet('产品型号清单')

    headers2 = ['序号', '产品型号', '支座编号范围', '数量', '生产日期', '检验标准']
    for col, h in enumerate(headers2, 1):
        ws2.cell(row=1, column=col, value=h)
    style_header(ws2, 1, len(headers2))

    samples2 = [
        [1, 'LNR600-Ⅱ', 'RA124017957-RA124017971', 15, '2024年11月', 'GB 20688.3-2006 / JG/T 118-2018'],
        [2, 'LNR700-Ⅱ', 'RA124017972-RA124018033', 62, '2024年11月', 'GB 20688.3-2006 / JG/T 118-2018'],
        [3, 'LNR700-Ⅱ-R', 'RA124018034-RA124018035', 2, '2024年11月', 'GB 20688.3-2006 / JG/T 118-2018'],
        [4, 'LRB700-Ⅱ', 'RA124018036-RA124018093', 58, '2024年11月', 'GB 20688.3-2006 / JG/T 118-2018'],
        [5, 'LRB700-Ⅱ-R', 'RA124018094-RA124018100', 7, '2024年11月', 'GB 20688.3-2006 / JG/T 118-2018'],
        [6, 'LNR835-Ⅱ', 'RA124018101-RA124018103', 3, '2024年11月', 'GB 20688.3-2006 / JG/T 118-2018'],
    ]
    for r, data in enumerate(samples2, 2):
        for c, val in enumerate(data, 1):
            ws2.cell(row=r, column=c, value=val)

    style_data_rows(ws2, 2, len(samples2) + 2, len(headers2))
    auto_width(ws2, len(headers2))
    ws2.column_dimensions['B'].width = 16
    ws2.column_dimensions['C'].width = 30
    ws2.column_dimensions['F'].width = 35

    # ===== Sheet 3: 力学性能检测数据 =====
    ws3 = wb.create_sheet('力学性能检测数据')

    headers3 = [
        '产品型号',
        '支座有效直径(mm)', '支座高度(mm)', '橡胶层总厚(mm)',
        '压缩应力(MPa)',
        '设计压缩刚度(KN/mm)',
        '水平等效刚度(KN/mm)',
        '屈服后刚度(KN/mm)',
        '屈服力(KN)',
        '等效阻尼比(%)',
    ]
    for col, h in enumerate(headers3, 1):
        ws3.cell(row=1, column=col, value=h)
    style_header(ws3, 1, len(headers3))

    samples3 = [
        ['LNR600-Ⅱ', 620, 208, 111, 12, 1800, 0.84, '/', '/', '/'],
        ['LNR700-Ⅱ', 720, 246, 129, 12, 2100, 0.99, '/', '/', '/'],
        ['LNR700-Ⅱ-R', 720, 246, 129, 12, 2100, 1.14, '/', '/', '/'],
        ['LRB700-Ⅱ', 720, 246, 129, 12, 2500, 1.66, 0.96, 90, '/'],
        ['LRB700-Ⅱ-R', 720, 246, 129, 12, 2500, 1.81, 1.11, 90, '/'],
        ['LNR835-Ⅱ', 855, 291, 153, 12, 2500, 1.34, '/', '/', '/'],
    ]
    for r, data in enumerate(samples3, 2):
        for c, val in enumerate(data, 1):
            ws3.cell(row=r, column=c, value=val)

    style_data_rows(ws3, 2, len(samples3) + 2, len(headers3))
    auto_width(ws3, len(headers3))
    ws3.column_dimensions['A'].width = 16

    # ===== Sheet 4: 外观尺寸检测数据 =====
    ws4 = wb.create_sheet('外观尺寸检测数据')

    headers4 = [
        '产品型号',
        '检测1-气泡(合格/不合格)',
        '检测2-杂质(合格/不合格)',
        '检测3-缺陷(合格/不合格)',
        '检测4-凹凸不平(合格/不合格)',
        '检测5-胶料粘结不良(合格/不合格)',
        '检测6-裂纹(合格/不合格)',
        '检测7-钢板外露(合格/不合格)',
        '检测8-产品标识(合格/不合格)',
        '检测9-平面尺寸偏差(合格/不合格)',
    ]
    for col, h in enumerate(headers4, 1):
        ws4.cell(row=1, column=col, value=h)
    style_header(ws4, 1, len(headers4))

    samples4 = [
        ['LNR600-Ⅱ'] + ['合格'] * 9,
        ['LNR700-Ⅱ'] + ['合格'] * 9,
        ['LNR700-Ⅱ-R'] + ['合格'] * 9,
        ['LRB700-Ⅱ'] + ['合格'] * 9,
        ['LRB700-Ⅱ-R'] + ['合格'] * 9,
        ['LNR835-Ⅱ'] + ['合格'] * 9,
    ]
    for r, data in enumerate(samples4, 2):
        for c, val in enumerate(data, 1):
            ws4.cell(row=r, column=c, value=val)

    style_data_rows(ws4, 2, len(samples4) + 2, len(headers4))
    auto_width(ws4, len(headers4))
    ws4.column_dimensions['A'].width = 16

    # Add note sheet
    ws5 = wb.create_sheet('填写说明')
    notes = [
        ['填写说明'],
        [''],
        ['1. 本模板用于生成隔震支座出厂检验报告'],
        ['2. Sheet"项目基本信息"：填写报告封面信息，每份报告填一行'],
        ['3. Sheet"产品型号清单"：填写本次报告包含的所有产品型号，每个型号一行'],
        ['4. Sheet"力学性能检测数据"：每个型号对应一行力学性能检测结果'],
        ['5. Sheet"外观尺寸检测数据"：每个型号对应一行外观检测结果'],
        ['6. 所有Sheet中的产品型号必须保持一致，程序将按型号名称进行数据关联'],
        ['7. 填写完成后保存，在软件中点击"导入Excel"即可'],
        ['8. 斜杠"/"表示该型号不需要填写该项（如天然橡胶支座无屈服力）'],
    ]
    for r, row_data in enumerate(notes, 1):
        for c, val in enumerate(row_data, 1):
            ws5.cell(row=r, column=c, value=val)
    ws5.column_dimensions['A'].width = 80

    # Save
    wb.save(output_path)
    print(f'Excel template saved to: {output_path}')
    return output_path


if __name__ == '__main__':
    _base = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    output = os.path.join(_base, 'products', 'isolation_bearing', 'excel_template.xlsx')
    create_excel_template(output)

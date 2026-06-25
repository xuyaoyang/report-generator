"""Read and parse the Excel parameter template."""
import os
import tempfile
from openpyxl import load_workbook


def read_excel_data(excel_path):
    """
    Read all sheets from the Excel template.
    Returns a structured dict:
    {
        'project_info': {...},
        'product_list': [{...}, {...}, ...],
        'mechanical_data': [{...}, {...}, ...],
        'visual_data': [{...}, {...}, ...],
    }
    """
    if not os.path.exists(excel_path):
        raise FileNotFoundError(f'Excel file not found: {excel_path}')

    recalculated_copy = None
    load_path = excel_path
    if _has_formulas(excel_path):
        recalculated_copy = _make_recalculated_copy(excel_path)
        if recalculated_copy:
            load_path = recalculated_copy

    wb = load_workbook(load_path, data_only=True)

    result = {}

    # Sheet 1: 项目基本信息
    if '项目基本信息' in wb.sheetnames:
        ws = wb['项目基本信息']
        headers = [cell.value for cell in ws[1]]
        values = [cell.value for cell in ws[2]]
        result['project_info'] = dict(zip(headers, values))

    # Sheet 2: 产品型号清单
    if '产品型号清单' in wb.sheetnames:
        ws = wb['产品型号清单']
        headers = [cell.value for cell in ws[1]]
        products = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is not None:  # skip empty rows
                products.append(dict(zip(headers, row)))
        result['product_list'] = products

    # Sheet 3: 力学性能检测数据
    if '力学性能检测数据' in wb.sheetnames:
        ws = wb['力学性能检测数据']
        headers = [cell.value for cell in ws[1]]
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is not None:
                data.append(dict(zip(headers, row)))
        result['mechanical_data'] = data

    # Sheet 4: 外观尺寸检测数据
    if '外观尺寸检测数据' in wb.sheetnames:
        ws = wb['外观尺寸检测数据']
        headers = [cell.value for cell in ws[1]]
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is not None:
                data.append(dict(zip(headers, row)))
        result['visual_data'] = data

    # Sheet 5: 产品材料需求 (optional)
    if '产品材料需求' in wb.sheetnames:
        ws = wb['产品材料需求']
        headers = [cell.value for cell in ws[1]]
        material_reqs = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is not None:
                material_reqs.append(dict(zip(headers, row)))
        result['material_requirements'] = material_reqs

    wb.close()
    if recalculated_copy:
        try:
            os.remove(recalculated_copy)
        except OSError:
            pass

    # Validate
    _validate_data(result)

    return result


def _has_formulas(excel_path):
    try:
        wb = load_workbook(excel_path, data_only=False, read_only=True)
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                if any(cell.data_type == 'f' for cell in row):
                    wb.close()
                    return True
        wb.close()
    except Exception:
        return False
    return False


def _make_recalculated_copy(excel_path):
    """Use local Excel to calculate formulas into a temporary copy.

    openpyxl reads cached formula results but does not calculate formulas.
    This keeps the user's original workbook untouched.
    """
    try:
        import pythoncom
        import win32com.client
    except Exception:
        return None

    tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    tmp_path = tmp.name
    tmp.close()

    excel = None
    workbook = None
    try:
        pythoncom.CoInitialize()
        excel = win32com.client.DispatchEx('Excel.Application')
        excel.Visible = False
        excel.DisplayAlerts = False
        workbook = excel.Workbooks.Open(
            os.path.abspath(excel_path),
            UpdateLinks=0,
            ReadOnly=True,
        )
        excel.CalculateFullRebuild()
        workbook.SaveCopyAs(tmp_path)
        return tmp_path
    except Exception:
        try:
            os.remove(tmp_path)
        except OSError:
            pass
        return None
    finally:
        if workbook is not None:
            workbook.Close(False)
        if excel is not None:
            excel.Quit()
        try:
            pythoncom.CoUninitialize()
        except Exception:
            pass


def _validate_data(data):
    """Basic validation of read data."""
    if not data.get('project_info') or not data['project_info'].get('项目名称'):
        raise ValueError('项目基本信息中缺少"项目名称"')
    if not data.get('product_list'):
        raise ValueError('产品型号清单为空，请至少填写一个产品型号')


def get_model_count(data):
    """Return the number of product models in the data."""
    return len(data.get('product_list', []))


def get_model_names(data):
    """Return list of product model names."""
    return [p['产品型号'] for p in data.get('product_list', [])]

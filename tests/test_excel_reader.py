import tempfile
import unittest
from unittest import mock

from openpyxl import Workbook

import core.excel_reader as excel_reader


class ExcelReaderTest(unittest.TestCase):
    def test_formula_recalculate_failure_blocks_import(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = f'{tmp}/formula.xlsx'
            wb = Workbook()
            ws = wb.active
            ws.title = '项目基本信息'
            ws.append(['项目名称'])
            ws.append(['测试项目'])
            ws['C1'] = '=1+1'
            wb.create_sheet('产品型号清单').append(['产品型号'])
            wb['产品型号清单'].append(['LRB'])
            wb.save(path)

            with mock.patch.object(
                    excel_reader, '_make_recalculated_copy',
                    return_value=(None, 'Excel unavailable')):
                with self.assertRaises(excel_reader.ExcelRecalculateError):
                    excel_reader.read_excel_data(path)


if __name__ == '__main__':
    unittest.main()

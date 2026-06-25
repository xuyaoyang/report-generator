"""Create the input workbook for friction-pendulum embedded-parts reports."""
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def _write_sheet(workbook, name, headers, rows, widths=None):
    sheet = workbook.create_sheet(name)
    for col, header in enumerate(headers, 1):
        cell = sheet.cell(1, col, header)
        cell.font = Font(name='Microsoft YaHei', bold=True)
        cell.fill = PatternFill('solid', fgColor='B3D9FF')
        cell.alignment = Alignment(horizontal='center', vertical='center')
    for row_index, row in enumerate(rows, 2):
        for col, value in enumerate(row, 1):
            cell = sheet.cell(row_index, col, value)
            cell.font = Font(name='Microsoft YaHei')
            cell.alignment = Alignment(horizontal='center', vertical='center')
    for col in range(1, len(headers) + 1):
        sheet.column_dimensions[get_column_letter(col)].width = (
            widths[col - 1] if widths else 20)
    sheet.freeze_panes = 'A2'
    return sheet


def create(output_path):
    workbook = Workbook()
    workbook.remove(workbook.active)
    _write_sheet(
        workbook, '项目基本信息',
        ['项目名称', '产品供应商', '制造地址', '联系电话', '制造日期',
         '检验员', '审核', '签发'],
        [['西藏震宇减震科技有限公司',
          '四川融海运通抗震科技有限责任公司',
          '四川省成都市新津区普兴街道清云南路256号',
          '18180838170', '2026年05月', '王正', '刘志建', '徐铫阳']],
        [42, 34, 42, 18, 18, 14, 14, 14])
    _write_sheet(
        workbook, '产品型号清单',
        ['支座规格', '支座数量', '生产批号', '生产日期',
         '预埋板边长(mm)', '预埋板厚度(mm)', '螺栓孔位置(mm)',
         '组件直径(mm)', '组件长度(mm)', '组件数量倍数', '钢板牌号'],
        [['FPS-II-3000×350-4.11', 22, '202605', '2026年5月',
          '', '', '', '', '', '', ''],
         ['FPS-II-4050×350-4.11', 18, '202605', '2026年5月',
          '', '', '', '', '', '', '']],
        [28, 16, 18, 18, 18, 18, 18, 16, 16, 16, 16])
    notes = workbook.create_sheet('填写说明')
    lines = [
        '填写说明',
        '1. 支座规格填写到摩擦摆支座主型号，不用填写-2/-3/-4/-5零件后缀。',
        '2. 程序会自动展开为上预埋板、下预埋板、上预埋组件、下预埋组件。',
        '3. 上/下预埋板数量等于支座数量；上/下预埋组件数量等于支座数量×4。',
        '4. 当前标准参数已内置FPS-II-3000和FPS-II-4050，新增规格需补充standard_connections.json。',
        '5. 表格中的边长、厚度、孔位、组件尺寸和数量倍数都可以手动填写；留空则按标准参数自动带出。',
        '6. 钢板材质单按厚度和Q235B自动匹配最新有效证书；如钢板牌号填写其它值，则按填写值匹配。',
    ]
    for row, line in enumerate(lines, 1):
        notes.cell(row, 1, line)
    notes.column_dimensions['A'].width = 100
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    workbook.save(output_path)


if __name__ == '__main__':
    create(os.path.join(os.path.dirname(os.path.abspath(__file__)),
                        'excel_template.xlsx'))

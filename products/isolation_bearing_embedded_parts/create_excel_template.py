"""Create the input workbook for isolation-bearing embedded-parts reports."""
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def _write_sheet(workbook, name, headers, rows, widths=None):
    sheet = workbook.create_sheet(name)
    for col, header in enumerate(headers, 1):
        cell = sheet.cell(1, col, header)
        cell.font = Font(name='Microsoft YaHei', bold=True)
        cell.fill = PatternFill('solid', fgColor='B3D9FF')
        cell.alignment = Alignment(horizontal='center', vertical='center')
    for row_index, row in enumerate(rows, 2):
        for col, value in enumerate(row, 1):
            cell = sheet.cell(row_index, col, value)
            cell.font = Font(name='Microsoft YaHei')
            cell.alignment = Alignment(horizontal='center', vertical='center')
    for col in range(1, len(headers) + 1):
        sheet.column_dimensions[get_column_letter(col)].width = (
            widths[col - 1] if widths else 20)
    sheet.freeze_panes = 'A2'
    return sheet


def create(output_path):
    workbook = Workbook()
    workbook.remove(workbook.active)
    _write_sheet(
        workbook, '\u9879\u76ee\u57fa\u672c\u4fe1\u606f',
        ['\u9879\u76ee\u540d\u79f0', '\u4ea7\u54c1\u4f9b\u5e94\u5546', '\u5236\u9020\u5730\u5740', '\u8054\u7cfb\u7535\u8bdd', '\u5236\u9020\u65e5\u671f',
         '\u68c0\u9a8c\u5458', '\u5ba1\u6838', '\u7b7e\u53d1'],
        [['\u897f\u660c\u9ad8\u67a7\u57ce\u4e2d\u6751\u68da\u6237\u533a\u6539\u9020\u9879\u76ee',
          '\u56db\u5ddd\u878d\u6d77\u8fd0\u901a\u6297\u9707\u79d1\u6280\u6709\u9650\u8d23\u4efb\u516c\u53f8',
          '\u56db\u5ddd\u7701\u6210\u90fd\u5e02\u65b0\u6d25\u533a\u666e\u5174\u8857\u9053\u6e05\u4e91\u5357\u8def256\u53f7',
          '18180838170', '2026\u5e745\u6708', '\u4f55\u6c38\u7434', '\u8463\u4fca\u8c6a', '\u5218\u5fd7\u5efa']],
        [42, 34, 42, 18, 18, 14, 14, 14])
    _write_sheet(
        workbook, '\u4ea7\u54c1\u578b\u53f7\u6e05\u5355',
        ['\u652f\u5ea7\u89c4\u683c', '\u652f\u5ea7\u6570\u91cf', '\u751f\u4ea7\u6279\u53f7', '\u751f\u4ea7\u65e5\u671f',
         '\u951a\u7b4b\u76f4\u5f84(mm)', '\u951a\u7b4b\u957f\u5ea6(mm)', '\u951a\u7b4b\u6570\u91cf\u500d\u6570',
         '\u5957\u7b52\u5916\u5f84(mm)', '\u5957\u7b52\u957f\u5ea6(mm)', '\u94a2\u677f\u8fb9\u957f(mm)',
         '\u94a2\u677f\u539a\u5ea6(mm)', '\u94a2\u7b4b\u724c\u53f7', '\u94a2\u677f\u724c\u53f7'],
        [['LNR500-II', 13, '202605', '2026\u5e745\u6708', '', '', '', '', '', '', '', '', ''],
         ['LNR700-II', 22, '202605', '2026\u5e745\u6708', '', '', '', '', '', '', '', '', ''],
         ['LNR900-II', 3, '202605', '2026\u5e745\u6708', '', '', '', '', '', '', '', '', '']],
        [22, 16, 18, 18, 16, 16, 16, 16, 16, 16, 16, 16, 16])
    notes = workbook.create_sheet('\u586b\u5199\u8bf4\u660e')
    lines = [
        '\u586b\u5199\u8bf4\u660e',
        '1. \u5f53\u524d\u6309II\u578b\u652f\u5ea7\u6807\u51c6\u8fde\u63a5\u53c2\u6570\u81ea\u52a8\u8ba1\u7b97\u3002',
        '2. \u652f\u5ea7\u89c4\u683c\u4e2d\u9700\u5305\u542b400~1600\u7684\u516c\u79f0\u76f4\u5f84\uff0c\u4f8b\u5982LNR700-II\u3002',
        '3. \u6bcf\u5957\u652f\u5ea7\u63091\u4ef6\u9884\u57cb\u94a2\u677f\u8ba1\u7b97\uff0c\u951a\u7b4b\u6309\u6807\u51c6\u8868\u768416/24/32/40\u500d\u8ba1\u7b97\u3002',
        '4. \u6750\u8d28\u5355\u6839\u636e\u94a2\u7b4b\u76f4\u5f84\u6216\u94a2\u677f\u539a\u5ea6\u3001\u6750\u8d28\u724c\u53f7\u81ea\u52a8\u5339\u914d\u6700\u65b0\u6709\u6548\u8bc1\u4e66\u3002',
        '5. \u951a\u7b4b\u76f4\u5f84/\u957f\u5ea6\u3001\u5957\u7b52\u5c3a\u5bf8\u3001\u94a2\u677f\u8fb9\u957f/\u539a\u5ea6\u7b49\u53c2\u6570\u53ef\u76f4\u63a5\u586b\u5199\uff1b\u7559\u7a7a\u5219\u6309\u6807\u51c6\u8868\u81ea\u52a8\u5e26\u51fa\u3002',
    ]
    for row, line in enumerate(lines, 1):
        notes.cell(row, 1, line)
    notes.column_dimensions['A'].width = 95
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    workbook.save(output_path)


if __name__ == '__main__':
    create(os.path.join(os.path.dirname(os.path.abspath(__file__)), 'excel_template.xlsx'))

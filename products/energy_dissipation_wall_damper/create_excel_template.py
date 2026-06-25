"""Create the input workbook for wall-board damper reports."""
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def _write_sheet(workbook, name, headers, rows, widths=None):
    sheet = workbook.create_sheet(name)
    for col, header in enumerate(headers, 1):
        cell = sheet.cell(1, col, header)
        cell.font = Font(name='Microsoft YaHei', bold=True)
        cell.fill = PatternFill('solid', fgColor='B3D9FF')
        cell.alignment = Alignment(horizontal='center', vertical='center')
    for row_index, row in enumerate(rows, 2):
        for col, value in enumerate(row, 1):
            cell = sheet.cell(row_index, col, value)
            cell.font = Font(name='Microsoft YaHei')
            cell.alignment = Alignment(horizontal='center', vertical='center')
    for col in range(1, len(headers) + 1):
        sheet.column_dimensions[get_column_letter(col)].width = (
            widths[col - 1] if widths else 20)
    sheet.freeze_panes = 'A2'
    return sheet


def create(output_path):
    workbook = Workbook()
    workbook.remove(workbook.active)

    _write_sheet(
        workbook,
        '项目基本信息',
        ['项目名称', '产品供应商', '制造地址', '联系电话', '制造日期',
         '检验员', '审核', '签发'],
        [['眉山天府新区幼儿园建设项目EPC（贵平幼儿园）',
          '四川融海运通抗震科技有限责任公司',
          '四川省成都市新津区普兴街道清云南路256号',
          '18180838170', '2026年6月', '01', '', '']],
        [44, 34, 42, 18, 16, 12, 12, 12],
    )

    _write_sheet(
        workbook,
        '产品型号清单',
        ['产品型号', '数量', '生产日期', '检验标准', '强度等级'],
        [['QB-1800-3360', 6, '2026年6月', 'JG/T209-2012', 'C35'],
         ['QB-1800-2560', 21, '2026年6月', 'JG/T209-2012', 'C35'],
         ['QB-1800-1960', 3, '2026年6月', 'JG/T209-2012', 'C35']],
        [24, 12, 16, 18, 14],
    )

    notes = workbook.create_sheet('填写说明')
    lines = [
        '填写说明',
        '1. 产品型号填写墙板阻尼器规格，如 QB-1800-3360。',
        '2. 数量填写数字即可，报告中会自动显示为“×件”。',
        '3. 生产日期用于合格证；制造日期用于封面报告日期。',
        '4. 检验标准、强度等级可按项目要求调整。',
        '5. 材质单图片在软件右侧“材质单图片”中勾选，生成时会追加到报告末尾。',
    ]
    for row, line in enumerate(lines, 1):
        notes.cell(row, 1, line)
    notes.column_dimensions['A'].width = 100

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    workbook.save(output_path)


if __name__ == '__main__':
    create(os.path.join(os.path.dirname(os.path.abspath(__file__)),
                        'excel_template.xlsx'))
